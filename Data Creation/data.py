import random
from datetime import datetime
from openpyxl import load_workbook

def get_value(cell):
	return cell.value

random.seed(datetime.now())

# Set number of profiles and projects to generate for randomized dataset
n_cand = 1000
n_role = 5000

# Input filenames to input randomized data
prof_fn = 'candidate_data.csv'
proj_fn = 'project_data.csv'

# Load sample data workbook
wb = load_workbook(filename='Dataset.xlsx')
ws = wb['Sheet1']

# Create dictionaries to store potential candidate and role attributes
candidate_vars = dict()
role_vars = dict()

# Populate the candidate dictionary with potential attributes
for col in ws.iter_cols(min_col=1, max_col=9, min_row=2):
	candidate_vars[col[0].value] = [i for i in map(get_value, col[1:]) if i is not None]

candidate_keys = candidate_vars.keys()

# Populate the role dictionary with potential attributes
for col in ws.iter_cols(min_col=11, max_col=18, min_row=2):
	role_vars[col[0].value] = [i for i in map(get_value, col[1:]) if i is not None]

role_keys = role_vars.keys()

# Randomly populate candidate attributes from candidate_data.csv
with open(prof_fn, 'w+') as profiles:
	line = 'UID,'
	for key in candidate_keys:
		line += key + ','
	line = line[:-1]
	profiles.write(line)

	for uid in range(n_cand):
		line = str(uid) + ','
		for key in candidate_keys:
			line += str(random.choice(candidate_vars[key])) + ','

		profiles.write('\n' + line[:-1])

# Randomly populate project attributes from project_data.csv
with open(proj_fn, 'w+') as projects:
	line = 'UID,'
	for key in role_keys:
		line += key + ','
	line = line[:-1]
	projects.write(line)

	for uid in range(n_role):
		line = str(uid) + ','

		start_date_idx = None
		start_date = None
		end_date = None

		for key in role_keys:
			elem = random.choice(role_vars[key])
			if key == 'Start Date':
				start_date_idx = role_vars[key].index(elem)
				start_date = elem
			elif key == 'End Date':
				elem = random.choice(role_vars[key][start_date_idx:])
				end_date = elem
			
			line += str(elem) + ','

		assert start_date < end_date
		projects.write('\n' + line[:-1])

