import random
from datetime import datetime
from openpyxl import load_workbook

def get_value(cell):
	return cell.value

random.seed(datetime.now())

# Set number of profiles and projects to generate for randomized dataset
n_cand = 1000
n_role = 5000

# Input filenames to input randomized data
prof_fn = 'Data/candidate_data.csv'
proj_fn = 'Data/project_data.csv'

# Load sample data workbook
wb = load_workbook(filename='Data/Dataset.xlsx')
ws = wb['Sheet1']

# Create dictionaries to store potential candidate and role attributes
candidate_vars = dict()
role_vars = dict()

# Populate the candidate dictionary with potential attributes
for col in ws.iter_cols(min_col=1, max_col=7, min_row=2):
	candidate_vars[col[0].value] = [i for i in map(get_value, col[1:]) if i is not None]

# candidate_keys = candidate_vars.keys()
candidate_keys = ["Service", "Talent Segment", "Level", "Skills", "Coming Available", "Local Office", "Local Preference"]

candidate_vars["Skills"] = {"IT Operations": ["ML", "Java", "Cloud", "Python", "Oracle", "Excel", "PowerPoint"], \
					   "Security": ["ML", "Java", "Cloud", "Python", "Oracle", "Excel", "PowerPoint"], \
					   "Design": ["PhotoShop", "Adobe", "React.js", "Excel", "PowerPoint"], \
					   "Legal": ["Bar", "Excel", "PowerPoint", "Oral communication"], \
					   "Human Resources": ["Excel", "PowerPoint", "Workday", "Oral communication"], \
					   "Account Leadership": ["ML", "Cloud", "Excel", "PowerPoint", "Leadership", "Oral communication"], \
					   "Finance": ["SAP", "Excel", "PowerPoint", "QuickBooks"], \
					   "Engineering": ["Python", "Java", "MATLAB", "Excel", "PowerPoint"], \
					   "Strategy": ["Excel", "PowerPoint", "Critical thinking"], \
					   "Business": ["Salesforce", "Excel", "PowerPoint"]}

# Populate the role dictionary with potential attributes
for col in ws.iter_cols(min_col=9, max_col=15, min_row=2):
	role_vars[col[0].value] = [i for i in map(get_value, col[1:]) if i is not None]

# role_keys = role_vars.keys()
role_keys = ["Start Date", "End Date", "Client", "Project Name", "Required Skills", "Location", "Local Requirement"]
role_vars["Required Skills"] = {"IT Operations": ["ML", "Java", "Cloud", "Python", "Oracle", "Excel", "PowerPoint"], \
					   "Security": ["ML", "Java", "Cloud", "Python", "Oracle", "Excel", "PowerPoint"], \
					   "Design": ["PhotoShop", "Adobe", "React.js", "Excel", "PowerPoint"], \
					   "Legal": ["Bar", "Excel", "PowerPoint", "Oral communication"], \
					   "Human Resources": ["Excel", "PowerPoint", "Workday", "Oral communication"], \
					   "Account Leadership": ["ML", "Cloud", "Excel", "PowerPoint", "Leadership", "Oral communication"], \
					   "Finance": ["SAP", "Excel", "PowerPoint", "QuickBooks"], \
					   "Engineering": ["Python", "Java", "MATLAB", "Excel", "PowerPoint"], \
					   "Strategy": ["Excel", "PowerPoint", "Critical thinking"], \
					   "Business": ["Salesforce", "Excel", "PowerPoint"]}

# Randomly populate candidate attributes from candidate_data.csv
with open(prof_fn, 'w+') as profiles:
	line = 'UID,'
	for key in candidate_keys:
		line += key + ','
	line = line[:-1]
	profiles.write(line)

	for uid in range(n_cand):
		talent_segment = None
		line = str(uid) + ','
		for key in candidate_keys:
			if key == "Talent Segment":
				talent_segment = random.choice(candidate_vars[key])
				line += str(talent_segment) + ','
			elif key == "Skills":
				n_skills = random.randint(1,4)
				skills = ""
				for i in range(n_skills):
					skills += str(random.choice(candidate_vars["Skills"][talent_segment]))+"/"
				line += str(skills) + ','
			else:
				line += str(random.choice(candidate_vars[key])) + ','

		profiles.write('\n' + line[:-1])

# Randomly populate project attributes from project_data.csv
with open(proj_fn, 'w+') as projects:
	line = 'UID,'
	for key in role_keys:
		line += key + ','
	projects.write(line + 'Level')

	for uid in range(n_role):
		project = None
		line = str(uid) + ','

		start_date_idx = None
		start_date = None
		end_date = None

		for key in role_keys:
			if uid == 0:
				print(key)
				print(line)

			if key == 'Start Date':
				elem = random.choice(role_vars[key])
				start_date_idx = role_vars[key].index(elem)
				start_date = elem
				line += str(elem) + ','
			elif key == 'End Date':
				elem = random.choice(role_vars[key][start_date_idx:])
				end_date = elem
				line += str(elem) + ','
			elif key == "Project Name":
				project = random.choice(role_vars[key])
				line += str(project) + " Project" + ','
				if uid == 0:
					print(project, line)
			elif key == "Required Skills":
				n_skills = random.randint(1,3)
				skills = ""
				for i in range(n_skills):
					skills += str(random.choice(role_vars["Required Skills"][project]))+"/"
				line += str(skills) + ','
			else:
				line += str(random.choice(role_vars[key])) + ','
				if uid == 0:
					print("else", line)

		# add level range
		lev = random.randint(2,11)
		line += str(lev-1) + '-' + str(lev)

		assert start_date < end_date
		projects.write('\n' + line)


